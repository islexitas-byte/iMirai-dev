import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def merge_and_style(ws, row, start_col, span, text, thin_border, fill="E0E0E0"):
    end_col = start_col + span - 1

    ws.merge_cells(
        start_row=row,
        start_column=start_col,
        end_row=row,
        end_column=end_col
    )

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", start_color=fill)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

        if col == start_col:
            cell.value = text
            cell.font = Font(name="Optima", size=11, bold=True)


def create_week_cells(ws, row, start_col, weeks,thin_border):
    for i in range(weeks):
        cell = ws.cell(row=row, column=start_col + i)
        cell.value = ""  # <-- NO week numbers
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

def apply_borders(ws, start_row, end_row, start_col, end_col,thin_border):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border

def save_project_plans(project_plan):
    try:
        project_plan = eval(project_plan)
        given_plan = pd.read_html(project_plan['sheet1'])
        given_plan = given_plan[0]
        given_plan['%'] = 0
        given_plan.fillna('',inplace=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "Progress"

        optima_font = Font(name="Optima", size=11)
        bold_optima = Font(name="Optima", size=11, bold=True)

        # Make header bold + light grey
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        header_font = Font(bold=True)
        wrap_align_right = Alignment(horizontal="right",)
        wrap_align_left = Alignment(horizontal="left",)
        wrap_align_center = Alignment(horizontal="center",)

        header_columns = list(given_plan.columns)
        # Header
        HEADER_ROWS_START = 1
        HEADER_ROWS_END = 2
        HEADER_COLS_END = 8  # columns 1 to 8

        for col in range(1, HEADER_COLS_END + 1):
            col_letter = get_column_letter(col)

            # Merge row 1 & 2 vertically
            ws.merge_cells(
                start_row=HEADER_ROWS_START,
                start_column=col,
                end_row=HEADER_ROWS_END,
                end_column=col
            )

            cell = ws.cell(row=1, column=col)
            cell.value = header_columns[col - 1]

            cell.font = bold_optima
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for cell in ws[2]:
            cell.border = thin_border

        # Sample data
        DATA_START_ROW = 3
        data = [tuple(v[1]) for v in given_plan.iterrows()]

        for r_idx, row in enumerate(data, start=DATA_START_ROW):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Styles
        bold_font = Font(bold=True)
        light_grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Apply rule: if S.No is NOT empty → bold + grey
        for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=len(header_columns)):
            sno_cell = row[0]
            if sno_cell.value not in (None, ""):
                for idx,cell in enumerate(row):
                    if idx==2:
                        # cell.font = bold_font
                        cell.alignment = wrap_align_left
                        cell.border = thin_border
                        cell.font = bold_optima
                    elif idx==0:
                        cell.alignment = wrap_align_right
                        cell.border = thin_border
                        cell.font = optima_font
                    else:
                        cell.alignment = wrap_align_center
                        cell.border = thin_border
                        cell.font = optima_font
                        
            else:
                for idx,cell in enumerate(row):
                    if idx==2:
                        cell.alignment = wrap_align_left
                        cell.border = thin_border
                        cell.font = optima_font
                        
                    elif idx==0:
                        cell.alignment = wrap_align_right
                        cell.border = thin_border
                        cell.font = optima_font
                        
                    else:
                        # print('bbb')
                        cell.alignment = wrap_align_center
                        cell.border = thin_border
                        cell.font = optima_font
                        
        year = "Year 2026"

        months = project_plan['Months and Weeks']

        year_row = 1
        month_row = 2
        week_row = 3
        start_col = 9  # Column D

        current_col = start_col
        total_weeks = sum(int(w) for _, w in months)
        ws.row_dimensions[month_row].height = 15
        ws.row_dimensions[week_row].height = 14
        for col in range(start_col, start_col + total_weeks):
            ws.column_dimensions[get_column_letter(col)].width = 2.36
        for col,val in enumerate([4.50,7.82,63.36,20.00,10.55,17.36,16.73,16.73]):
            ws.column_dimensions[get_column_letter(col+1)].width = val
        # YEAR (merged across all weeks)
        merge_and_style(
            ws,
            row=year_row,
            start_col=start_col,
            span=total_weeks,
            text=year,
            thin_border=thin_border,
            fill="D6D6D6"
        )

        # MONTHS + WEEKS
        for month, weeks in months:
            # Month row
            weeks = int(weeks)
            merge_and_style(
                ws,
                row=month_row,
                start_col=current_col,
                span=weeks,
                text=month,
                thin_border=thin_border,
                fill="E0E0E0"
            )

            # Week row
            create_week_cells(
                ws,
                row=week_row,
                start_col=current_col,
                weeks=weeks,
                thin_border=thin_border
            )
            
            apply_borders(
            ws,
            start_row=1,          # Year row
            end_row=len(given_plan)+2,     # Week row
            start_col=start_col,
            end_col=start_col + total_weeks - 1,
            thin_border=thin_border
            )

            current_col += weeks

        # Green-only color scale for %
        green_scale = ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="E8F5E9",   # light green
            end_type="num",
            end_value=100,
            end_color="1B5E20"      # dark green
        )

        last_data_row = DATA_START_ROW + len(given_plan) - 1
        ws.conditional_formatting.add(
            f"B{DATA_START_ROW}:B{last_data_row}",
            green_scale
        )
        wb.save(project_plan['file_name'])
        return 'Done'
    except Exception as e:
        print(e)
        return None

